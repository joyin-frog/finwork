#!/usr/bin/env python3
"""
生成 Excel 预览增强测试 fixture。
覆盖:公式、货币/千分位/百分比/小数格式、加粗+上边框合计行、表头加粗、合并单元格、文本+数字混排。

输出: tests/fixtures/excel-preview-enhance.xlsx
"""
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要 openpyxl:pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "tests", "fixtures", "excel-preview-enhance.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "财务报表"

# ── 表头(加粗 + 合并单元格) ──────────────────────────────────────────────────
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "2024年度收支汇总表"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal="center")

# ── 列头(加粗) ────────────────────────────────────────────────────────────────
headers = ["项目", "Q1金额", "Q2金额", "Q3金额", "合计"]
header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="E8F0FE")
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

# ── 数据行 ──────────────────────────────────────────────────────────────────
data = [
    ("产品收入",  120000.50,  98765.00,  150000.00),
    ("服务收入",   45000.00,  52000.00,   48000.00),
    ("其他收入",    8000.00,   7500.00,    9200.00),
    ("销售费用",  -30000.00, -28000.00,  -35000.00),
    ("管理费用",  -20000.00, -22000.00,  -21000.00),
]

currency_fmt = "¥#,##0.00"
int_fmt = "#,##0"
pct_fmt = "0.0%"

for row_idx, (name, q1, q2, q3) in enumerate(data, start=3):
    ws.cell(row=row_idx, column=1, value=name)
    for col, val in enumerate([q1, q2, q3], start=2):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.number_format = currency_fmt
    # 合计公式
    sum_col = 5
    q1_addr = f"B{row_idx}"
    q3_addr = f"D{row_idx}"
    total_cell = ws.cell(row=row_idx, column=sum_col)
    total_cell.value = f"=SUM({q1_addr}:{q3_addr})"
    total_cell.number_format = currency_fmt

# ── 汇率行(百分比示例) ───────────────────────────────────────────────────────
pct_row = len(data) + 3
ws.cell(row=pct_row, column=1, value="增长率")
pct_vals = [0.123, -0.045, 0.089]
for col, v in enumerate(pct_vals, 2):
    c = ws.cell(row=pct_row, column=col, value=v)
    c.number_format = pct_fmt
# 公式:A1*B1 示例
avg_cell = ws.cell(row=pct_row, column=5)
avg_cell.value = "=B8*C8"
avg_cell.number_format = "0.00%"

# ── 合计行(加粗 + 上边框) ────────────────────────────────────────────────────
total_row = len(data) + 3 + 1
thick_top = Border(top=Side(style="medium"))
bold_font = Font(bold=True)
ws.cell(row=total_row, column=1, value="净收入合计").font = bold_font
ws.cell(row=total_row, column=1).border = thick_top

for col in range(2, 6):
    row_letter = get_column_letter(col)
    start = 3
    end = 3 + len(data) - 1
    c = ws.cell(row=total_row, column=col)
    c.value = f"=SUM({row_letter}{start}:{row_letter}{end})"
    c.number_format = currency_fmt
    c.font = bold_font
    c.border = thick_top

# ── 整数千分位示例行 ────────────────────────────────────────────────────────
int_row = total_row + 2
ws.cell(row=int_row, column=1, value="员工人数")
for col, val in enumerate([120, 130, 125], 2):
    c = ws.cell(row=int_row, column=col, value=val)
    c.number_format = int_fmt
total_emp = ws.cell(row=int_row, column=5)
total_emp.value = f"=SUM(B{int_row}:D{int_row})"
total_emp.number_format = int_fmt

# ── 负数示例 ─────────────────────────────────────────────────────────────────
neg_row = int_row + 1
ws.cell(row=neg_row, column=1, value="亏损项目")
neg_cell = ws.cell(row=neg_row, column=2, value=-99999.99)
neg_cell.number_format = "¥#,##0.00;[Red]-¥#,##0.00"

# ── 设置列宽 ─────────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 18
for col in ["B", "C", "D", "E"]:
    ws.column_dimensions[col].width = 16

wb.save(OUT)
print(f"生成完成: {OUT}")
