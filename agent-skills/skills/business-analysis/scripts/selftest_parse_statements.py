#!/usr/bin/env python3
"""parse_statements.py 的确定性自测(无 pytest 依赖:plain assert + 退出码)。

构造合成会小企工作簿——**故意把列位摆成真实金蝶导出的样子**(资产负债表科目名在第0列、
行次在第1列;利润表行次在第2列、本年累计在第4列)——验证解析器靠「探测列」而非写死列号
取对数。由 tests/business-analysis-script.test.ts 经 venv python 跑进 CI(脚本的确定性守护)。
"""
import os
import sys
import tempfile
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parse_statements import parse_workbook  # noqa: E402


def build_workbook(path):
    wb = openpyxl.Workbook()
    # 资产负债表:两栏式。资产侧 列0=科目名/列1=行次/列2=期末;负债侧 列4=科目名/列5=行次/列6=期末
    bs = wb.active
    bs.title = "资产负债表"
    bs.append(["资产负债表"])
    bs.append(["编制单位:测试公司(单位:元)"])
    bs.append(["资产", "行次", "期末余额", "期初余额", "负债和所有者权益", "行次", "期末余额", "期初余额"])
    bs.append(["货币资金", 1, 500, 400, "短期借款", 31, 3000, 2800])
    bs.append(["应收账款", 4, 2000, 1800, "应付账款", 33, 1500, 1400])
    bs.append(["存货", 9, 1000, 900, "流动负债合计", 41, 6000, 5500])
    bs.append(["流动资产合计", 15, 5000, 4500, "负债合计", None, 6000, 5500])
    bs.append(["资产总计", None, 10000, 9000, "所有者权益合计", None, 4000, 3500])

    # 利润表:行次在第2列、本月在第3列、本年累计在第4列、上年同期在第5列(本年累计应被优先取)
    is_ = wb.create_sheet("利润表")
    is_.append(["利润表"])
    is_.append(["项目", None, "行次", "本月金额", "本年累计金额", "上年同期"])
    is_.append(["一、营业收入", None, 1, 700, 8000, 6000])
    is_.append(["减:营业成本", None, 2, 450, 5000, 3800])
    is_.append(["销售费用", None, 11, 10, 100, 90])
    is_.append(["管理费用", None, 14, 60, 600, 550])
    is_.append(["研究费用", None, 17, 40, 400, 380])
    is_.append(["财务费用", None, 18, 20, 200, 180])
    is_.append(["四、净利润", None, 32, 50, 500, 420])

    # 现金流量表:按科目名取净额
    cf = wb.create_sheet("现金流量表")
    cf.append(["现金流量表"])
    cf.append(["项目", "行次", "本期金额"])
    cf.append(["经营活动产生的现金流量净额", 7, 1900])
    cf.append(["投资活动产生的现金流量净额", 13, -100])
    cf.append(["筹资活动产生的现金流量净额", 20, 200])
    cf.append(["现金及现金等价物净增加额", 21, 2000])

    # T3 敏感 sheet(客户实名)——解析器不应读它,这里只作存在性干扰
    ar = wb.create_sheet("应收账款明细")
    ar.append(["单位", "归属", "金额"])
    ar.append(["某客户", "某集团", 99999])

    wb.save(path)


def main():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "fin.xlsx")
        build_workbook(p)
        out = parse_workbook(p)

    bs, is_, cf = out["balanceSheet"], out["incomeStatement"], out["cashFlow"]
    prior = (is_ or {}).get("prior", {})
    fails = []

    def chk(cond, msg):
        if not cond:
            fails.append(msg)

    chk(out["unit"] == "元", f"unit 应=元,实际 {out['unit']}")
    chk(bs and bs["cash"] == 500 and bs["receivables"] == 2000 and bs["inventory"] == 1000 and bs["currentAssets"] == 5000,
        f"BS 资产侧行次取数错(列位探测): {bs}")
    chk(bs and bs["shortTermBorrowing"] == 3000 and bs["payables"] == 1500 and bs["currentLiabilities"] == 6000,
        f"BS 负债侧行次取数错: {bs}")
    chk(bs and bs["totalAssets"] == 10000 and bs["totalLiabilities"] == 6000 and bs["equity"] == 4000,
        f"BS 合计(科目名兜底)取数错: {bs}")
    chk(is_ and is_["revenue"] == 8000 and is_["cost"] == 5000 and is_["netProfit"] == 500,
        f"IS 应取本年累计(非本月): {is_}")
    chk(is_ and is_["sellingExpense"] == 100 and is_["adminExpense"] == 600 and is_["rdExpense"] == 400 and is_["financeExpense"] == 200,
        f"IS 费用行次取数错: {is_}")
    chk(prior.get("revenue") == 6000 and prior.get("netProfit") == 420, f"IS 上年同期取数错: {prior}")
    chk(cf and cf["operatingCashFlow"] == 1900 and cf["netCashIncrease"] == 2000, f"CF 科目名取数错: {cf}")

    if fails:
        print("FAIL:\n" + "\n".join(fails))
        sys.exit(1)
    print("PASS: parse_statements 确定性解析 8 项断言通过(列位探测 + 行次/科目名取数)")


if __name__ == "__main__":
    main()
