#!/usr/bin/env python3
"""会小企 01/02/03 财报三表确定性解析(business-analysis skill 固定脚本)。

为什么是固定脚本而非"agent 现写 python":同样输入 → 同样输出、可回算(红线 2)。
按"行次/科目名"取数,空单元格 = 0(红线 4)。**T3 敏感 sheet(应收/进项客户实名明细)
不在解析范围、不读入(红线 7)。**

比 lib 里写死列号的解析更稳:**自动探测「行次」列与「期末余额/本年累计金额」列**,
兼容各家金蝶导出列位漂移(实测金蝶导出行次在第 1 列、非第 0 列)。

用法:
    python3 parse_statements.py <财报.xlsx>
输出:stdout 一段 canonical JSON,字段与 generate_business_analysis 工具入参对齐。
"""
import sys
import json
import re
import openpyxl
from openpyxl.utils import get_column_letter

# ── 行次 → canonical 字段(会小企 2025 国标行次)──────────────────────────
BS_BY_LINE = {1: "cash", 4: "receivables", 9: "inventory", 15: "currentAssets",
              31: "shortTermBorrowing", 33: "payables", 41: "currentLiabilities"}
# 资产负债表常见两栏式没有“行次”，所有关键科目都必须支持科目名提取。
BS_LABELS = {
    "cash": ["货币资金"],
    "receivables": ["应收账款"],
    "inventory": ["存货"],
    "currentAssets": ["流动资产合计"],
    "shortTermBorrowing": ["短期借款"],
    "payables": ["应付账款"],
    "currentLiabilities": ["流动负债合计"],
    "totalAssets": ["资产总计", "资产合计"],
    "totalLiabilities": ["负债合计"],
    "equity": ["所有者权益合计", "净资产合计", "股东权益合计", "所有者权益(或股东权益)合计"],
}
IS_BY_LINE = {1: "revenue", 2: "cost", 11: "sellingExpense", 14: "adminExpense",
              18: "financeExpense", 32: "netProfit"}
# 会小企02表行17是管理费用下的“其中：研究费用”，已经包含在行14管理费用中，
# 不能再次映射成并列研发费用。只有报表明确单列“研发费用”时才提取 rdExpense。
IS_STANDALONE = {"rdExpense": ["研发费用"]}
CF_TOTALS = {
    "operatingCashFlow": ["经营活动产生的现金流量净额", "经营活动现金流量净额"],
    "investingCashFlow": ["投资活动产生的现金流量净额"],
    "financingCashFlow": ["筹资活动产生的现金流量净额"],
    "netCashIncrease": ["现金及现金等价物净增加额", "现金净增加额"],
}


def _num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").strip())
        except ValueError:
            return None
    return None


def _grid(ws, max_r=140, max_c=16):
    rows = min(ws.max_row or 1, max_r)
    cols = min(ws.max_column or 1, max_c)
    # read_only worksheet 的随机 ws.cell() 每次都会从 XML 开头扫描，宽表上会退化为
    # 数十秒甚至分钟级；一次顺序迭代能保持固定上限并避免受虚假 max_column 影响。
    return [list(row) for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols, values_only=True)]


def _detect_unit(grid):
    for row in grid[:8]:
        for cell in row:
            if isinstance(cell, str) and "万元" in cell:
                return "万元"
    return "元"


# 取值列偏好:本年累计优先(年度比率口径),其次期末余额/本期/本月
_VALUE_PREF = ["本年累计金额", "期末余额", "期末数", "本期金额", "本月金额"]


def _header_groups(grid, prior=False):
    """每个「行次」表头 → (行次列, 取值列)。两栏式资产负债表会返回两组。"""
    groups = []
    for row in grid[:10]:
        for c, cell in enumerate(row):
            if not (isinstance(cell, str) and cell.strip() == "行次"):
                continue
            if prior:
                val_col = next((cc for cc in range(c + 1, len(row))
                                if isinstance(row[cc], str)
                                and any(header in row[cc] for header in ("上年同期", "年初数", "期初数", "期初余额"))), None)
            else:
                val_col, best_rank = None, len(_VALUE_PREF)
                for cc in range(c + 1, len(row)):
                    h = row[cc]
                    if not isinstance(h, str):
                        continue
                    for rank, kw in enumerate(_VALUE_PREF):
                        if kw in h and rank < best_rank:
                            val_col, best_rank = cc, rank
            if val_col is not None:
                groups.append((c, val_col))
    if groups:
        return groups
    # 无“行次”的两栏式资产负债表：把值列自身作为标签锚点，_label_left 会取左侧科目。
    prior_headers = ("年初数", "期初数", "期初余额", "上年同期")
    current_headers = tuple(_VALUE_PREF)
    wanted = prior_headers if prior else current_headers
    for row in grid[:10]:
        for c, cell in enumerate(row):
            if isinstance(cell, str) and any(header in cell for header in wanted):
                groups.append((c, c))
    return groups


def _label_left(row, lc):
    for cc in range(lc - 1, -1, -1):
        if 0 <= cc < len(row) and isinstance(row[cc], str) and row[cc].strip():
            return row[cc].strip()
    return ""


def _by_line(grid, line_map, groups):
    res = {}
    for (lc, vc) in groups:
        if lc == vc:
            continue
        for row in grid:
            if lc >= len(row):
                continue
            ln = _num(row[lc])
            if ln is None:
                continue
            field = line_map.get(int(ln))
            if field and vc < len(row):
                n = _num(row[vc])
                if n is not None:
                    res[field] = n
    return res


def _locators_by_line(ws, grid, line_map, groups, prefix):
    res = {}
    for (lc, vc) in groups:
        if lc == vc:
            continue
        for row_index, row in enumerate(grid, start=1):
            if lc >= len(row):
                continue
            ln = _num(row[lc])
            if ln is None:
                continue
            field = line_map.get(int(ln))
            if field and vc < len(row) and _num(row[vc]) is not None:
                res[f"{prefix}.{field}"] = {
                    "sheet": ws.title,
                    "range": f"{get_column_letter(vc + 1)}{row_index}",
                }
    return res


def _label_matches(label, names):
    # 锚定匹配:整体相等,或以总计名结尾且不是「流动/非流动」小计——避免「流动资产合计」误命中「资产合计」
    normalized_label = re.sub(r"\s+", "", label).rstrip(":：")
    return any(
        normalized_label == re.sub(r"\s+", "", nm).rstrip(":：")
        or (
            normalized_label.endswith(re.sub(r"\s+", "", nm).rstrip(":："))
            and not normalized_label.startswith(("流动", "非流动"))
        )
        for nm in names
    )


def _by_label(grid, label_map, groups):
    res = {}
    for (lc, vc) in groups:
        for row in grid:
            label = _label_left(row, lc)
            if not label or vc >= len(row):
                continue
            for field, names in label_map.items():
                if field not in res and _label_matches(label, names):
                    n = _num(row[vc])
                    if n is not None:
                        res[field] = n
    return res


def _locators_by_label(ws, grid, label_map, groups, prefix):
    res = {}
    for (lc, vc) in groups:
        for row_index, row in enumerate(grid, start=1):
            label = _label_left(row, lc)
            if not label or vc >= len(row):
                continue
            for field, names in label_map.items():
                key = f"{prefix}.{field}"
                if key not in res and _label_matches(label, names) and _num(row[vc]) is not None:
                    res[key] = {
                        "sheet": ws.title,
                        "range": f"{get_column_letter(vc + 1)}{row_index}",
                    }
    return res


def _zero_fill(d, fields):
    return {f: d.get(f, 0.0) for f in fields}


BS_FIELDS = ["cash", "receivables", "inventory", "currentAssets", "totalAssets",
             "shortTermBorrowing", "payables", "currentLiabilities", "totalLiabilities", "equity"]
IS_FIELDS = ["revenue", "cost", "sellingExpense", "adminExpense", "rdExpense", "financeExpense", "netProfit"]
CF_FIELDS = ["operatingCashFlow", "investingCashFlow", "financingCashFlow", "netCashIncrease"]


def _find_sheet(wb, *keywords):
    for name in wb.sheetnames:
        if any(k in name for k in keywords):
            return wb[name]
    return None


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {"balanceSheet": None, "incomeStatement": None, "cashFlow": None, "unit": "元", "sourceCells": {}}

    bs = _find_sheet(wb, "资产负债")
    if bs is not None:
        g = _grid(bs)
        groups = _header_groups(g)
        d = {**_by_line(g, BS_BY_LINE, groups), **_by_label(g, BS_LABELS, groups)}
        out["balanceSheet"] = _zero_fill(d, BS_FIELDS)
        out["sourceCells"].update(_locators_by_line(bs, g, BS_BY_LINE, groups, "balanceSheet"))
        out["sourceCells"].update(_locators_by_label(bs, g, BS_LABELS, groups, "balanceSheet"))
        prior_groups = _header_groups(g, prior=True)
        prior = {**_by_line(g, BS_BY_LINE, prior_groups), **_by_label(g, BS_LABELS, prior_groups)}
        if any(prior.values()):
            prior_fields = ["cash", "receivables", "inventory", "currentAssets", "totalAssets",
                            "currentLiabilities", "totalLiabilities", "equity"]
            out["balanceSheet"]["prior"] = _zero_fill(prior, prior_fields)
            out["sourceCells"].update(_locators_by_line(bs, g, BS_BY_LINE, prior_groups, "balanceSheet.prior"))
            out["sourceCells"].update(_locators_by_label(bs, g, BS_LABELS, prior_groups, "balanceSheet.prior"))
        out["unit"] = _detect_unit(g)

    is_sheet = _find_sheet(wb, "利润", "损益")
    if is_sheet is not None:
        g = _grid(is_sheet)
        groups = _header_groups(g)
        cur = {**_by_line(g, IS_BY_LINE, groups), **_by_label(g, IS_STANDALONE, groups)}
        out["incomeStatement"] = _zero_fill(cur, IS_FIELDS)
        out["sourceCells"].update(_locators_by_line(is_sheet, g, IS_BY_LINE, groups, "incomeStatement"))
        out["sourceCells"].update(_locators_by_label(is_sheet, g, IS_STANDALONE, groups, "incomeStatement"))
        prior_groups = _header_groups(g, prior=True)
        if prior_groups:
            prior = {**_by_line(g, IS_BY_LINE, prior_groups), **_by_label(g, IS_STANDALONE, prior_groups)}
            if any(prior.values()):
                out["incomeStatement"]["prior"] = _zero_fill(prior, IS_FIELDS)
                out["sourceCells"].update(_locators_by_line(is_sheet, g, IS_BY_LINE, prior_groups, "incomeStatement.prior"))
                out["sourceCells"].update(_locators_by_label(is_sheet, g, IS_STANDALONE, prior_groups, "incomeStatement.prior"))

    cf = _find_sheet(wb, "现金流")
    if cf is not None:
        g = _grid(cf)
        groups = _header_groups(g)
        out["cashFlow"] = _zero_fill(_by_label(g, CF_TOTALS, groups), CF_FIELDS)
        out["sourceCells"].update(_locators_by_label(cf, g, CF_TOTALS, groups, "cashFlow"))

    wb.close()
    return out


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "用法: parse_statements.py <财报.xlsx>"}, ensure_ascii=False))
        sys.exit(2)
    try:
        print(json.dumps(parse_workbook(sys.argv[1]), ensure_ascii=False))
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"error": f"解析失败: {e}"}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
